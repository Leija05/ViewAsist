from dotenv import load_dotenv
load_dotenv()

from fastapi import FastAPI, APIRouter, HTTPException, Request, UploadFile, File, Depends, Response
from fastapi.responses import StreamingResponse, FileResponse
from starlette.background import BackgroundTask
from starlette.middleware.cors import CORSMiddleware
from starlette.concurrency import run_in_threadpool
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
import os
import logging
import asyncio
import bcrypt
import jwt
import io
import csv
import json
import secrets
import socket
import tempfile
import ipaddress
import traceback
import unicodedata
from glob import glob
from pathlib import Path
from urllib.parse import urlparse
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any, Literal
from datetime import datetime, timezone, timedelta

try:
    from zk import ZK
except Exception:
    ZK = None

socket.setdefaulttimeout(10)


ROOT_DIR = Path(__file__).parent

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7  # 7 days for persistent session

def get_jwt_secret() -> str:
    return os.environ["JWT_SECRET"]

def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    hashed = bcrypt.hashpw(password.encode("utf-8"), salt)
    return hashed.decode("utf-8")

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return bcrypt.checkpw(plain_password.encode("utf-8"), hashed_password.encode("utf-8"))

def create_access_token(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "exp": datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES),
        "type": "access"
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {
        "sub": user_id,
        "exp": datetime.now(timezone.utc) + timedelta(days=30),
        "type": "refresh"
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def get_xlrd_module():
    try:
        import xlrd
        return xlrd
    except ModuleNotFoundError as exc:
        raise HTTPException(
            status_code=503,
            detail='Falta dependencia opcional "xlrd". Ejecuta: pip install xlrd==2.0.1',
        ) from exc

def get_openpyxl_module():
    try:
        import openpyxl
        return openpyxl
    except ModuleNotFoundError as exc:
        raise HTTPException(
            status_code=503,
            detail='Falta dependencia opcional "openpyxl". Ejecuta: pip install openpyxl',
        ) from exc

class ExcelSheetAdapter:
    def __init__(self, sheet_name: str, nrows: int, ncols: int, getter):
        self.sheet_name = sheet_name
        self.nrows = nrows
        self.ncols = ncols
        self._getter = getter

    def cell_value(self, row_idx: int, col_idx: int):
        return self._getter(row_idx, col_idx)

class ExcelWorkbookAdapter:
    def __init__(self, sheet_map: Dict[str, ExcelSheetAdapter]):
        self._sheet_map = sheet_map

    def sheet_names(self):
        return list(self._sheet_map.keys())

    def sheet_by_name(self, name: str):
        return self._sheet_map[name]

def load_excel_workbook(filename: str, content: bytes) -> ExcelWorkbookAdapter:
    extension = Path(filename or "").suffix.lower()

    if extension == ".xlsx":
        openpyxl = get_openpyxl_module()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        sheets = {}
        for ws in wb.worksheets:
            sheets[ws.title] = ExcelSheetAdapter(
                sheet_name=ws.title,
                nrows=ws.max_row or 0,
                ncols=ws.max_column or 0,
                getter=lambda row_idx, col_idx, _ws=ws: (_ws.cell(row=row_idx + 1, column=col_idx + 1).value or "")
            )
        return ExcelWorkbookAdapter(sheets)

    xlrd = get_xlrd_module()
    wb = xlrd.open_workbook(file_contents=content)
    sheets = {}
    for sheet_name in wb.sheet_names():
        sheet = wb.sheet_by_name(sheet_name)
        sheets[sheet_name] = ExcelSheetAdapter(
            sheet_name=sheet_name,
            nrows=sheet.nrows,
            ncols=sheet.ncols,
            getter=lambda row_idx, col_idx, _sheet=sheet: _sheet.cell_value(row_idx, col_idx)
        )
    return ExcelWorkbookAdapter(sheets)


def _normalize_sheet_name(value: str) -> str:
    return str(value or "").strip().lower().replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")


def _find_header_row(sheet: ExcelSheetAdapter, required_labels: List[str], max_scan_rows: int = 15) -> Optional[int]:
    required = [_normalize_sheet_name(label) for label in required_labels]
    for row_idx in range(min(max_scan_rows, sheet.nrows)):
        values = [
            _normalize_sheet_name(sheet.cell_value(row_idx, col_idx))
            for col_idx in range(sheet.ncols)
        ]
        if all(any(req in val for val in values) for req in required):
            return row_idx
    return None


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(float(value))
    except Exception:
        return default

def get_reportlab_modules():
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        return {
            "colors": colors,
            "letter": letter,
            "landscape": landscape,
            "SimpleDocTemplate": SimpleDocTemplate,
            "Table": Table,
            "TableStyle": TableStyle,
            "Paragraph": Paragraph,
            "Spacer": Spacer,
            "getSampleStyleSheet": getSampleStyleSheet,
            "ParagraphStyle": ParagraphStyle,
        }
    except ModuleNotFoundError as exc:
        raise HTTPException(
            status_code=503,
            detail='Falta dependencia opcional "reportlab". Ejecuta: pip install reportlab',
        ) from exc

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="No autenticado")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Tipo de token inválido")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="Usuario no encontrado")
        user["_id"] = str(user["_id"])
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")

# Create the main app
app = FastAPI(title="Sistema de Control de Asistencia")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Pydantic Models
class LoginRequest(BaseModel):
    email: EmailStr
    password: str
    remember_me: bool = True

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    role: str

class EmployeeBase(BaseModel):
    employee_id: str
    name: str
    department: str

class AttendanceRecord(BaseModel):
    employee_id: str
    name: str
    department: str
    date: str
    entry_time: Optional[str] = None
    exit_time: Optional[str] = None
    delay_minutes: float = 0
    early_departure_minutes: float = 0
    absence_minutes: float = 0
    status: str  # "presente", "retardo", "falta"

class SettingsUpdate(BaseModel):
    entry_time: str = "09:00"
    tolerance_minutes: int = 30
    work_hours: int = 9

class AttendanceRule(BaseModel):
    name: str
    expected_entry_time: str
    tolerance_minutes: int = 30

class ClockConfigUpdate(BaseModel):
    device_name: str = "Reloj Principal"
    ip: str
    port: int = 4370
    password: str = ""
    rules: List[AttendanceRule] = Field(default_factory=list)

class ClockSyncResponse(BaseModel):
    synced_records: int
    generated_report_id: Optional[str] = None
    message: str
    detected_records: Optional[int] = 0
    skipped_duplicates: Optional[int] = 0
    inserted_records: Optional[int] = 0

class ClockConnectionPayload(BaseModel):
    connected: bool


class ClockConfigImportPayload(BaseModel):
    device_name: str = "Reloj Principal"
    ip: str
    port: int = 4370
    password: str = ""
    rules: List[AttendanceRule] = Field(default_factory=list)


class AttSettingRow(BaseModel):
    numero: int
    entrada: str
    salida: str
    tiempo_extra: int = 0


class AttSettingsPayload(BaseModel):
    settings: List[AttSettingRow] = Field(default_factory=list)

class UsbScanPayload(BaseModel):
    mount_path: str


class ClockUserBase(BaseModel):
    user_id: str
    name: str
    department: str = "General"
    privilege: Literal["admin", "empleado"] = "empleado"
    password: str = ""
    card_number: str = ""
    fingerprint_registered: bool = False
    face_registered: bool = False
    vein_registered: bool = False
    work_schedule: str = "Turno General"
    enabled: bool = True

class ClockUserUpdate(BaseModel):
    name: Optional[str] = None
    department: Optional[str] = None
    privilege: Optional[Literal["admin", "empleado"]] = None
    password: Optional[str] = None
    card_number: Optional[str] = None
    fingerprint_registered: Optional[bool] = None
    face_registered: Optional[bool] = None
    vein_registered: Optional[bool] = None
    work_schedule: Optional[str] = None
    enabled: Optional[bool] = None

class VersionInfo(BaseModel):
    current_version: str
    latest_version: Optional[str] = None
    update_available: bool = False
    release_notes: Optional[str] = None
    download_url: Optional[str] = None


class ClockUserSyncPayload(BaseModel):
    sync_mode: Literal["wifi", "manual"] = "manual"


BONO_ENTRY_START = "09:00"
BONO_ENTRY_END = "09:30"
ADMIN_SCHEDULES = {
    "leonel puente": {"entry": "09:00", "exit": "18:00"},
    "anahi espinoza": {"entry": "09:00", "exit": "17:00"},
    "alejandro muñiz": {"entry": "11:00", "exit": "19:00"},
    "alejandro muniz": {"entry": "11:00", "exit": "19:00"},
}

# Auth Routes
@api_router.post("/auth/login")
async def login(request: LoginRequest, response: Response):
    email = request.email.lower()
    user = await db.users.find_one({"email": email})
    
    if not user or not verify_password(request.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    
    user_id = str(user["_id"])
    access_token = create_access_token(user_id, email)
    refresh_token = create_refresh_token(user_id)
    
    max_age = 604800 if request.remember_me else 3600  # 7 days or 1 hour
    
    response.set_cookie(
        key="access_token",
        value=access_token,
        httponly=True,
        secure=False,
        samesite="lax",
        max_age=max_age,
        path="/"
    )
    response.set_cookie(
        key="refresh_token",
        value=refresh_token,
        httponly=True,
        secure=False,
        samesite="lax",
        max_age=2592000,  # 30 days
        path="/"
    )
    
    return {
        "id": user_id,
        "email": user["email"],
        "name": user["name"],
        "role": user["role"]
    }

@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"message": "Sesión cerrada"}

@api_router.get("/auth/me")
async def get_me(request: Request):
    user = await get_current_user(request)
    return user

@api_router.post("/auth/refresh")
async def refresh_token(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="No hay token de refresh")
    
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Tipo de token inválido")
        
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="Usuario no encontrado")
        
        new_access_token = create_access_token(str(user["_id"]), user["email"])
        
        response.set_cookie(
            key="access_token",
            value=new_access_token,
            httponly=True,
            secure=False,
            samesite="lax",
            max_age=604800,
            path="/"
        )
        
        return {"message": "Token refrescado"}
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token de refresh expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")

# Excel Processing
@api_router.post("/upload/excel")
async def upload_excel(request: Request, file: UploadFile = File(...)):
    await get_current_user(request)
    
    if not file.filename.endswith(('.xls', '.xlsx')):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos Excel (.xls, .xlsx)")
    
    content = await file.read()
    
    try:
        # Get settings
        settings = await db.settings.find_one({}, {"_id": 0})
        if not settings:
            settings = {"entry_time": "09:00", "tolerance_minutes": 30, "work_hours": 9}
        
        entry_hour, entry_min = map(int, settings["entry_time"].split(":"))
        tolerance = settings["tolerance_minutes"]
        
        # Parse Excel
        workbook = load_excel_workbook(file.filename, content)
        
        all_data = {
            "filename": file.filename,
            "upload_date": datetime.now(timezone.utc).isoformat(),
            "sheets": {},
            "employees": [],
            "attendance_records": [],
            "statistics": {}
        }
        
        # Process Reporte de Excepciones sheet (main attendance data)
        if "Reporte de Excepciones" in workbook.sheet_names():
            sheet = workbook.sheet_by_name("Reporte de Excepciones")
            records = []
            
            for row_idx in range(4, sheet.nrows):
                try:
                    emp_id = str(sheet.cell_value(row_idx, 0)).strip()
                    if not emp_id or emp_id == "":
                        continue
                    
                    name = str(sheet.cell_value(row_idx, 1)).strip()
                    department = str(sheet.cell_value(row_idx, 2)).strip()
                    date = str(sheet.cell_value(row_idx, 3)).strip()
                    entry_time = str(sheet.cell_value(row_idx, 4)).strip()
                    exit_time = str(sheet.cell_value(row_idx, 5)).strip()
                    
                    delay_min = float(sheet.cell_value(row_idx, 8) or 0)
                    early_dep = float(sheet.cell_value(row_idx, 9) or 0)
                    absence_min = float(sheet.cell_value(row_idx, 10) or 0)
                    
                    # Determine status based on tolerance
                    if absence_min >= 540:  # Full day absence (9 hours)
                        status = "falta"
                    elif delay_min > tolerance:
                        status = "retardo"
                    elif delay_min > 0:
                        status = "retardo_tolerancia"
                    else:
                        status = "presente"
                    
                    record = {
                        "employee_id": emp_id,
                        "name": name,
                        "department": department,
                        "date": date,
                        "entry_time": entry_time if entry_time else None,
                        "exit_time": exit_time if exit_time else None,
                        "delay_minutes": delay_min,
                        "early_departure_minutes": early_dep,
                        "absence_minutes": absence_min,
                        "status": status
                    }
                    records.append(record)
                except Exception as e:
                    continue
            
            all_data["attendance_records"] = records
        
        # Process Reporte Estadístico sheet
        if "Reporte Estadístico" in workbook.sheet_names():
            sheet = workbook.sheet_by_name("Reporte Estadístico")
            employees = []
            
            for row_idx in range(4, sheet.nrows):
                try:
                    emp_id = str(sheet.cell_value(row_idx, 0)).strip()
                    if not emp_id or emp_id == "":
                        continue
                    
                    name = str(sheet.cell_value(row_idx, 1)).strip()
                    department = str(sheet.cell_value(row_idx, 2)).strip()
                    normal_hours = str(sheet.cell_value(row_idx, 3)).strip()
                    real_hours = str(sheet.cell_value(row_idx, 4)).strip()
                    delay_count = float(sheet.cell_value(row_idx, 5) or 0)
                    delay_minutes = float(sheet.cell_value(row_idx, 6) or 0)
                    early_count = float(sheet.cell_value(row_idx, 7) or 0)
                    early_minutes = float(sheet.cell_value(row_idx, 8) or 0)
                    absence_days = float(sheet.cell_value(row_idx, 13) or 0)
                    
                    employee = {
                        "employee_id": emp_id,
                        "name": name,
                        "department": department,
                        "normal_hours": normal_hours,
                        "real_hours": real_hours,
                        "delay_count": int(delay_count),
                        "delay_minutes": int(delay_minutes),
                        "early_departure_count": int(early_count),
                        "early_departure_minutes": int(early_minutes),
                        "absence_days": int(absence_days)
                    }
                    employees.append(employee)
                except Exception as e:
                    continue
            
            all_data["employees"] = employees
        
        # Extract raw sheet data for preview
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            sheet_data = []
            for row_idx in range(min(50, sheet.nrows)):  # Limit to 50 rows for preview
                row_data = []
                for col_idx in range(sheet.ncols):
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    row_data.append(str(cell_value) if cell_value else "")
                sheet_data.append(row_data)
            all_data["sheets"][sheet_name] = sheet_data
        
        # Calculate statistics
        total_employees = len(all_data["employees"])
        total_absences = sum(e["absence_days"] for e in all_data["employees"])
        total_delays = sum(e["delay_count"] for e in all_data["employees"])
        total_delay_minutes = sum(e["delay_minutes"] for e in all_data["employees"])
        
        # Find employees with most issues
        employees_sorted_by_absences = sorted(all_data["employees"], key=lambda x: x["absence_days"], reverse=True)
        employees_sorted_by_delays = sorted(all_data["employees"], key=lambda x: x["delay_count"], reverse=True)
        
        all_data["statistics"] = {
            "total_employees": total_employees,
            "total_absences": int(total_absences),
            "total_delays": int(total_delays),
            "total_delay_minutes": int(total_delay_minutes),
            "avg_absences_per_employee": round(total_absences / total_employees, 2) if total_employees > 0 else 0,
            "avg_delays_per_employee": round(total_delays / total_employees, 2) if total_employees > 0 else 0,
            "top_absent_employees": employees_sorted_by_absences[:5],
            "top_delayed_employees": employees_sorted_by_delays[:5]
        }
        
        # Save to database
        report_doc = {
            "filename": all_data["filename"],
            "upload_date": datetime.now(timezone.utc),
            "employees": all_data["employees"],
            "attendance_records": all_data["attendance_records"],
            "statistics": all_data["statistics"],
            "raw_content": content
        }
        
        result = await db.reports.insert_one(report_doc)
        all_data["report_id"] = str(result.inserted_id)
        
        # Don't return raw_content
        del all_data["sheets"]  # Remove raw sheets from response
        
        return all_data
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error procesando archivo: {str(e)}")


def _parse_standard_report_sections(filename: str, content: bytes) -> Dict[str, Any]:
    workbook = load_excel_workbook(filename, content)
    sections: Dict[str, List[Dict[str, Any]]] = {"estadistico": [], "turnos": [], "asistencia": [], "excepciones": []}
    employees_map: Dict[str, Dict[str, Any]] = {}

    for sheet_name in workbook.sheet_names():
        normalized_name = _normalize_sheet_name(sheet_name)
        sheet = workbook.sheet_by_name(sheet_name)
        header_row = _find_header_row(sheet, ["id", "nombre"], max_scan_rows=20)
        if header_row is None:
            continue

        headers = [str(sheet.cell_value(header_row, c) or "").strip() for c in range(sheet.ncols)]
        id_col = next((idx for idx, h in enumerate(headers) if "id" in _normalize_sheet_name(h) or _normalize_sheet_name(h) in {"no.", "no", "numero"}), 0)
        name_col = next((idx for idx, h in enumerate(headers) if "nombre" in _normalize_sheet_name(h)), 1)
        dept_col = next((idx for idx, h in enumerate(headers) if "depart" in _normalize_sheet_name(h) or "depto" in _normalize_sheet_name(h)), None)

        section_key = (
            "estadistico" if "estad" in normalized_name else
            "turnos" if "turno" in normalized_name else
            "asistencia" if "asist" in normalized_name else
            "excepciones" if "excep" in normalized_name else
            "asistencia"
        )

        for row_idx in range(header_row + 1, sheet.nrows):
            employee_id = str(sheet.cell_value(row_idx, id_col) or "").strip()
            if not employee_id:
                continue
            row_data: Dict[str, str] = {}
            for col_idx, column_name in enumerate(headers):
                key = column_name or f"col_{col_idx + 1}"
                row_data[key] = str(sheet.cell_value(row_idx, col_idx) or "").strip()
            sections[section_key].append(row_data)
            employees_map.setdefault(employee_id, {
                "employee_id": employee_id,
                "name": str(sheet.cell_value(row_idx, name_col) or "").strip(),
                "department": str(sheet.cell_value(row_idx, dept_col) or "").strip() if dept_col is not None else "General",
                "absences": 0,
                "delays": 0,
                "assistances": 0,
            })

    for row in sections["estadistico"]:
        emp_id = str(row.get("ID") or row.get("No.") or row.get("No") or row.get("No. empleado") or "").strip()
        if not emp_id or emp_id not in employees_map:
            continue
        employees_map[emp_id]["absences"] = _safe_int(row.get("Ausencias") or row.get("Faltas") or row.get("Días de ausencia"))
        employees_map[emp_id]["delays"] = _safe_int(row.get("Retardos") or row.get("No. retardos") or row.get("Retardo"))
        employees_map[emp_id]["assistances"] = max(0, _safe_int(row.get("Asistencias") or row.get("Días laborales"), 0))

    return {"employees": list(employees_map.values()), "sections": sections}


def _parse_attsettings(filename: str, content: bytes) -> Dict[str, Any]:
    workbook = load_excel_workbook(filename, content)
    horarios: List[Dict[str, Any]] = []
    turnos: List[Dict[str, Any]] = []
    for sheet_name in workbook.sheet_names():
        normalized_name = _normalize_sheet_name(sheet_name)
        sheet = workbook.sheet_by_name(sheet_name)
        for row_idx in range(sheet.nrows):
            values = [str(sheet.cell_value(row_idx, c) or "").strip() for c in range(sheet.ncols)]
            if not any(values):
                continue
            target = turnos if "turno" in normalized_name else horarios
            target.append({"row_index": row_idx, "values": values})
    return {"horarios": horarios, "turnos": turnos}


@api_router.post("/usb/scan")
async def scan_usb_files(payload: UsbScanPayload, request: Request):
    await get_current_user(request)
    base = Path(str(payload.mount_path or "").strip()).expanduser()
    if not str(base).strip() or not base.exists() or not base.is_dir():
        raise HTTPException(status_code=400, detail="Ruta USB inválida")

    standard_matches = sorted(glob(str(base / "**" / "*StandardReport.xls*"), recursive=True))
    att_matches = sorted(glob(str(base / "**" / "*AttSetting.xls*"), recursive=True))
    return {
        "status": "ok",
        "mount_path": str(base),
        "detected": {
            "standard_report": standard_matches[0] if standard_matches else None,
            "attsettings": att_matches[0] if att_matches else None,
        },
        "all_matches": {"standard_report": standard_matches, "attsettings": att_matches},
    }


@api_router.post("/usb/load")
async def load_usb_payload(
    request: Request,
    standard_report: UploadFile = File(None),
    attsettings: UploadFile = File(None),
):
    await get_current_user(request)
    if not standard_report and not attsettings:
        raise HTTPException(status_code=400, detail="Sube al menos *StandardReport.xls o *AttSetting.xls")

    payload: Dict[str, Any] = {"status": "ok", "source": "usb", "loaded_at": datetime.now(timezone.utc).isoformat()}
    if standard_report:
        standard_content = await standard_report.read()
        parsed_standard = _parse_standard_report_sections(standard_report.filename or "", standard_content)
        payload["standard_report"] = {"filename": standard_report.filename, **parsed_standard}
    if attsettings:
        settings_content = await attsettings.read()
        payload["attsettings"] = {"filename": attsettings.filename, "config": _parse_attsettings(attsettings.filename or "", settings_content)}
    return payload

@api_router.get("/reports")
async def get_reports(request: Request):
    await get_current_user(request)
    reports = await db.reports.find({}, {"raw_content": 0}).sort("upload_date", -1).to_list(100)
    for r in reports:
        r["_id"] = str(r["_id"])
        if isinstance(r.get("upload_date"), datetime):
            r["upload_date"] = r["upload_date"].isoformat()
    return reports

@api_router.get("/reports/{report_id}")
async def get_report(report_id: str, request: Request):
    await get_current_user(request)
    if report_id in {"csv", "export"}:
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        if not start_date or not end_date:
            raise HTTPException(status_code=400, detail="Debes enviar start_date y end_date en formato YYYY-MM-DD.")
        return await export_clock_events_csv(request=request, start_date=start_date, end_date=end_date)
    try:
        report = await db.reports.find_one({"_id": ObjectId(report_id)}, {"raw_content": 0})
        if not report:
            raise HTTPException(status_code=404, detail="Reporte no encontrado")
        report["_id"] = str(report["_id"])
        if isinstance(report.get("upload_date"), datetime):
            report["upload_date"] = report["upload_date"].isoformat()
        return report
    except Exception:
        raise HTTPException(status_code=404, detail="Reporte no encontrado")

@api_router.get("/reports/{report_id}/excel-preview")
async def get_excel_preview(report_id: str, request: Request):
    await get_current_user(request)
    try:
        report = await db.reports.find_one({"_id": ObjectId(report_id)})
        if not report or "raw_content" not in report:
            raise HTTPException(status_code=404, detail="Reporte no encontrado")
        
        content = report["raw_content"]
        workbook = load_excel_workbook(report.get("filename", ""), content)
        
        sheets_data = {}
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            sheet_data = []
            for row_idx in range(sheet.nrows):
                row_data = []
                for col_idx in range(sheet.ncols):
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    row_data.append(str(cell_value) if cell_value else "")
                sheet_data.append(row_data)
            sheets_data[sheet_name] = sheet_data
        
        return {"sheets": sheets_data, "filename": report.get("filename", "")}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/reports/{report_id}")
async def delete_report(report_id: str, request: Request):
    await get_current_user(request)
    try:
        result = await db.reports.delete_one({"_id": ObjectId(report_id)})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Reporte no encontrado")
        return {"message": "Reporte eliminado"}
    except Exception:
        raise HTTPException(status_code=404, detail="Reporte no encontrado")

# Settings
@api_router.get("/settings")
async def get_settings(request: Request):
    await get_current_user(request)
    settings = await db.settings.find_one({}, {"_id": 0})
    if not settings:
        settings = {"entry_time": "09:00", "tolerance_minutes": 30, "work_hours": 9}
        await db.settings.insert_one(settings)
    return settings

@api_router.put("/settings")
async def update_settings(settings: SettingsUpdate, request: Request):
    await get_current_user(request)
    await db.settings.update_one({}, {"$set": settings.model_dump()}, upsert=True)
    return settings.model_dump()

# Version & Updates
@api_router.get("/version")
async def get_version():
    current_version = os.environ.get("APP_VERSION", "1.0.0")
    github_repo = os.environ.get("GITHUB_REPO", "")
    
    return {
        "current_version": current_version,
        "github_repo": github_repo,
        "update_available": False,
        "latest_version": current_version,
        "release_notes": None,
        "download_url": None
    }

@api_router.post("/check-updates")
async def check_updates():
    """Check for updates from GitHub releases"""
    import requests
    
    current_version = os.environ.get("APP_VERSION", "1.0.0")
    github_repo = os.environ.get("GITHUB_REPO", "")
    
    if not github_repo:
        return {
            "current_version": current_version,
            "update_available": False,
            "message": "No hay repositorio de GitHub configurado"
        }
    
    try:
        # Format: owner/repo
        api_url = f"https://api.github.com/repos/{github_repo}/releases/latest"
        response = requests.get(api_url, timeout=10)
        
        if response.status_code == 200:
            release = response.json()
            latest_version = release.get("tag_name", "").lstrip("v")
            
            # Simple version comparison
            current_parts = [int(x) for x in current_version.split(".")]
            latest_parts = [int(x) for x in latest_version.split(".")]
            
            update_available = latest_parts > current_parts
            
            return {
                "current_version": current_version,
                "latest_version": latest_version,
                "update_available": update_available,
                "release_notes": release.get("body", ""),
                "download_url": release.get("html_url", ""),
                "published_at": release.get("published_at", "")
            }
        else:
            return {
                "current_version": current_version,
                "update_available": False,
                "message": "No se pudo obtener información de actualizaciones"
            }
    except Exception as e:
        return {
            "current_version": current_version,
            "update_available": False,
            "message": f"Error verificando actualizaciones: {str(e)}"
        }

# PDF Export
@api_router.get("/reports/{report_id}/pdf")
async def export_pdf(report_id: str, request: Request):
    await get_current_user(request)
    
    try:
        report = await db.reports.find_one({"_id": ObjectId(report_id)}, {"raw_content": 0})
        if not report:
            raise HTTPException(status_code=404, detail="Reporte no encontrado")
        
        reportlab = get_reportlab_modules()
        colors = reportlab["colors"]
        SimpleDocTemplate = reportlab["SimpleDocTemplate"]
        Table = reportlab["Table"]
        TableStyle = reportlab["TableStyle"]
        Paragraph = reportlab["Paragraph"]
        Spacer = reportlab["Spacer"]
        getSampleStyleSheet = reportlab["getSampleStyleSheet"]
        ParagraphStyle = reportlab["ParagraphStyle"]
        letter = reportlab["letter"]
        landscape = reportlab["landscape"]

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, spaceAfter=20)
        elements.append(Paragraph(f"Reporte de Asistencia - {report.get('filename', 'Sin nombre')}", title_style))
        elements.append(Spacer(1, 12))
        
        # Statistics
        stats = report.get("statistics", {})
        stats_text = f"""
        <b>Estadísticas Generales:</b><br/>
        Total de Empleados: {stats.get('total_employees', 0)}<br/>
        Total de Faltas: {stats.get('total_absences', 0)}<br/>
        Total de Retardos: {stats.get('total_delays', 0)}<br/>
        Minutos de Retardo Total: {stats.get('total_delay_minutes', 0)}<br/>
        """
        elements.append(Paragraph(stats_text, styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Employee Table
        elements.append(Paragraph("<b>Resumen por Empleado:</b>", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        table_data = [["ID", "Nombre", "Departamento", "Faltas", "Retardos", "Min. Retardo"]]
        for emp in report.get("employees", []):
            table_data.append([
                emp.get("employee_id", ""),
                emp.get("name", "")[:25],
                emp.get("department", "")[:15],
                str(emp.get("absence_days", 0)),
                str(emp.get("delay_count", 0)),
                str(emp.get("delay_minutes", 0))
            ])
        
        if len(table_data) > 1:
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"reporte_asistencia_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando PDF: {str(e)}")


@api_router.get("/reports/export")
@api_router.get("/reports/csv")
async def export_clock_events_csv(
    request: Request,
    start_date: str,
    end_date: str,
):
    await get_current_user(request)

    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Usa YYYY-MM-DD.")
    if end_dt < start_dt:
        raise HTTPException(status_code=400, detail="end_date no puede ser menor que start_date.")

    users = await db.clock_users.find({}, {"user_id": 1, "name": 1}).to_list(5000)
    user_name_by_id = {
        str(u.get("user_id", "")).strip(): str(u.get("name", "")).strip() or f"ID {u.get('user_id', '')}"
        for u in users
        if str(u.get("user_id", "")).strip()
    }

    daily_rows: Dict[tuple, Dict[str, Any]] = {}
    start_day = start_dt.date()
    end_day = end_dt.date()

    docs = await db.clock_events.find({}, {"events": 1}).sort("created_at", -1).to_list(5000)
    for doc in docs:
        for event in doc.get("events", []):
            raw_user_id = str(
                event.get("clock_user_id")
                or event.get("employee_id")
                or event.get("user_id")
                or ""
            ).strip()
            if not raw_user_id:
                continue

            raw_timestamp = event.get("timestamp")
            if isinstance(raw_timestamp, datetime):
                event_dt = raw_timestamp
            else:
                try:
                    event_dt = datetime.fromisoformat(str(raw_timestamp))
                except Exception:
                    continue

            event_day = event_dt.date()
            if event_day < start_day or event_day > end_day:
                continue

            key = (raw_user_id, event_day.isoformat())
            current = daily_rows.get(key)
            if not current:
                current = {
                    "employee_id": raw_user_id,
                    "name": user_name_by_id.get(raw_user_id) or event.get("employee_name") or f"ID Reloj: {raw_user_id}",
                    "date": event_day.isoformat(),
                    "entry": event_dt,
                    "exit": event_dt,
                    "delay_minutes": float(event.get("delay_minutes", 0) or 0),
                }
                daily_rows[key] = current
            else:
                if event_dt < current["entry"]:
                    current["entry"] = event_dt
                if event_dt > current["exit"]:
                    current["exit"] = event_dt
                current["delay_minutes"] = max(current["delay_minutes"], float(event.get("delay_minutes", 0) or 0))

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Nombre", "Fecha", "Entrada", "Salida", "Retardos"])
    for row in sorted(daily_rows.values(), key=lambda r: (r["date"], r["employee_id"])):
        writer.writerow([
            row["employee_id"],
            row["name"],
            row["date"],
            row["entry"].strftime("%H:%M:%S"),
            row["exit"].strftime("%H:%M:%S"),
            round(row["delay_minutes"], 2),
        ])

    filename = f"reporte_historico_{start_day.isoformat()}_{end_day.isoformat()}.csv"
    temp_file = tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False, encoding="utf-8-sig", newline="")
    try:
        temp_file.write(output.getvalue())
        temp_file_path = temp_file.name
    finally:
        temp_file.close()
        output.close()

    return FileResponse(
        path=temp_file_path,
        media_type="text/csv; charset=utf-8",
        filename=filename,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        background=BackgroundTask(lambda: os.path.exists(temp_file_path) and os.unlink(temp_file_path)),
    )

# Dashboard Stats
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(request: Request):
    await get_current_user(request)
    
    # Get latest report
    latest_report = await db.reports.find_one({}, {"raw_content": 0}, sort=[("upload_date", -1)])
    
    if not latest_report:
        return {
            "has_data": False,
            "message": "No hay reportes cargados"
        }
    
    stats = latest_report.get("statistics", {})
    employees = latest_report.get("employees", [])
    
    # Alerts for employees with high absences or delays
    alerts = []
    for emp in employees:
        if emp.get("absence_days", 0) >= 5:
            alerts.append({
                "type": "danger",
                "employee": emp["name"],
                "message": f"Tiene {emp['absence_days']} faltas"
            })
        elif emp.get("delay_count", 0) >= 5:
            alerts.append({
                "type": "warning", 
                "employee": emp["name"],
                "message": f"Tiene {emp['delay_count']} retardos"
            })
    
    return {
        "has_data": True,
        "report_id": str(latest_report["_id"]),
        "filename": latest_report.get("filename", ""),
        "upload_date": latest_report.get("upload_date").isoformat() if isinstance(latest_report.get("upload_date"), datetime) else latest_report.get("upload_date"),
        "statistics": stats,
        "employees": employees,
        "alerts": alerts[:10]  # Top 10 alerts
    }

# Employees
@api_router.get("/employees")
async def get_employees(request: Request):
    await get_current_user(request)
    latest_report = await db.reports.find_one({}, {"employees": 1}, sort=[("upload_date", -1)])
    if not latest_report:
        return []
    return latest_report.get("employees", [])

@api_router.get("/employees/{employee_id}/history")
async def get_employee_history(employee_id: str, request: Request):
    await get_current_user(request)
    
    # Get all reports and find records for this employee
    reports = await db.reports.find({}, {"attendance_records": 1, "upload_date": 1, "filename": 1}).to_list(100)
    
    history = []
    for report in reports:
        records = [r for r in report.get("attendance_records", []) if r.get("employee_id") == employee_id]
        if records:
            history.append({
                "report_date": report.get("upload_date").isoformat() if isinstance(report.get("upload_date"), datetime) else report.get("upload_date"),
                "filename": report.get("filename", ""),
                "records": records
            })
    
    return history

def _summarize_clock_records(records: List[Dict[str, Any]], settings: Dict[str, Any], user_lookup: Optional[Dict[str, str]] = None) -> Dict[str, Any]:
    def _normalize_name(value: str) -> str:
        normalized = unicodedata.normalize("NFKD", value or "")
        return "".join(ch for ch in normalized if not unicodedata.combining(ch)).strip().lower()

    user_lookup = user_lookup or {}
    grouped: Dict[tuple, List[datetime]] = {}

    for record in records:
        employee_id = str(record.get("employee_id", "")).strip()
        timestamp = record.get("timestamp")
        if not employee_id or not isinstance(timestamp, datetime):
            continue
        ts = timestamp if timestamp.tzinfo else timestamp.replace(tzinfo=timezone.utc)
        key = (employee_id, ts.date().isoformat())
        grouped.setdefault(key, []).append(ts)

    entry_hour, entry_minute = map(int, settings.get("entry_time", "09:00").split(":"))
    tolerance_minutes = int(settings.get("tolerance_minutes", 30))
    bono_start_h, bono_start_m = map(int, BONO_ENTRY_START.split(":"))
    bono_end_h, bono_end_m = map(int, BONO_ENTRY_END.split(":"))

    attendance_records = []
    employee_summary: Dict[str, Dict[str, Any]] = {}

    for (employee_id, date_str), punches in grouped.items():
        punches = sorted(punches)
        entry_time = punches[0]
        exit_time = punches[-1]

        employee_name = user_lookup.get(employee_id, f"Empleado {employee_id}")
        employee_name_key = _normalize_name(employee_name)
        custom_schedule = ADMIN_SCHEDULES.get(employee_name_key)
        schedule_entry = custom_schedule["entry"] if custom_schedule else f"{entry_hour:02d}:{entry_minute:02d}"
        schedule_entry_hour, schedule_entry_minute = map(int, schedule_entry.split(":"))

        scheduled = datetime.fromisoformat(
            f"{date_str}T{schedule_entry_hour:02d}:{schedule_entry_minute:02d}:00"
        ).replace(tzinfo=timezone.utc)
        delay_minutes = max((entry_time - scheduled).total_seconds() / 60.0, 0)
        status = "retardo" if delay_minutes > tolerance_minutes else "presente"
        bono_start_dt = datetime.fromisoformat(f"{date_str}T{bono_start_h:02d}:{bono_start_m:02d}:00").replace(tzinfo=timezone.utc)
        bono_end_dt = datetime.fromisoformat(f"{date_str}T{bono_end_h:02d}:{bono_end_m:02d}:00").replace(tzinfo=timezone.utc)
        bonus_aplica = bono_start_dt <= entry_time <= bono_end_dt

        attendance_records.append({
            "employee_id": employee_id,
            "name": employee_name,
            "department": "Reloj checador",
            "date": date_str,
            "entry_time": entry_time.strftime("%H:%M"),
            "exit_time": exit_time.strftime("%H:%M"),
            "delay_minutes": round(delay_minutes, 2),
            "early_departure_minutes": 0,
            "absence_minutes": 0,
            "status": status,
            "bonus_aplica": bonus_aplica,
            "bonus_horario": f"{BONO_ENTRY_START} a {BONO_ENTRY_END}",
            "admin_schedule": custom_schedule or None,
        })

        if employee_id not in employee_summary:
            employee_summary[employee_id] = {
                "employee_id": employee_id,
                "name": employee_name,
                "department": "Reloj checador",
                "absence_days": 0,
                "delay_count": 0,
                "delay_minutes": 0,
                "bonus_eligible_days": 0,
                "bonus_lost_days": 0,
            }

        if status == "retardo":
            employee_summary[employee_id]["delay_count"] += 1
            employee_summary[employee_id]["delay_minutes"] += round(delay_minutes, 2)
        if bonus_aplica:
            employee_summary[employee_id]["bonus_eligible_days"] += 1
        else:
            employee_summary[employee_id]["bonus_lost_days"] += 1

    employees = list(employee_summary.values())
    statistics = {
        "total_employees": len(employees),
        "total_absences": 0,
        "total_delays": sum(e["delay_count"] for e in employees),
        "total_delay_minutes": round(sum(e["delay_minutes"] for e in employees), 2),
        "bonus_window": f"{BONO_ENTRY_START} a {BONO_ENTRY_END}",
        "bonus_policy_note": "Después de ese horario ya no cuenta para bono. Las faltas se mantienen para control de vacaciones.",
    }

    return {
        "attendance_records": attendance_records,
        "employees": employees,
        "statistics": statistics,
    }


def _fetch_clock_attendance(config: Dict[str, Any]) -> Dict[str, Any]:
    conn = _connect_to_clock(config)
    try:
        conn.disable_device()
        users = conn.get_users() or []
        user_lookup = {str(u.user_id): (u.name or f"Empleado {u.user_id}") for u in users}
        attendance = conn.get_attendance() or []

        normalized = []
        for item in attendance:
            timestamp = getattr(item, "timestamp", None)
            user_id = str(getattr(item, "user_id", "")).strip()
            if not timestamp or not user_id:
                continue
            normalized.append({"employee_id": user_id, "timestamp": timestamp})

        return {"records": normalized, "users": user_lookup}
    except TimeoutError as exc:
        raise HTTPException(status_code=408, detail=f"Tiempo de espera agotado al leer asistencias del reloj: {exc}")
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Error al leer asistencias del reloj: {exc}")
    finally:
        try:
            conn.enable_device()
        except Exception:
            pass
        try:
            conn.disconnect()
        except Exception:
            pass


def _connect_to_clock(config: Dict[str, Any]):
    device_ip = str(config.get("ip", "") or "").strip()
    if not device_ip:
        raise HTTPException(status_code=400, detail="Ingresa la IP del reloj antes de conectar.")

    raw_port = config.get("port", 4370)
    try:
        device_port = int(raw_port)
    except Exception:
        raise HTTPException(status_code=400, detail="El puerto del reloj no es válido.")
    if device_port <= 0:
        raise HTTPException(status_code=400, detail="El puerto del reloj no es válido.")

    raw_password = str(config.get("password", "") or "").strip()
    if raw_password == "":
        raise HTTPException(status_code=400, detail="Ingresa la Contraseña/Comm Key del reloj antes de conectar.")

    try:
        from zk import ZK  # type: ignore
    except Exception as exc:
        raise HTTPException(
            status_code=503,
            detail=f'No se encontró la librería pyzk/zk: {exc}. Instala dependencias con: pip install pyzk'
        )

    try:
        device_password = int(raw_password)
    except Exception:
        raise HTTPException(status_code=400, detail="La Contraseña/Comm Key debe ser numérica.")

    zk_client = ZK(device_ip, port=device_port, timeout=10, password=device_password, force_udp=False, ommit_ping=True)
    try:
        return zk_client.connect()
    except Exception as exc:
        message = str(exc)
        if "Unauthenticated" in message:
            raise HTTPException(
                status_code=400,
                detail="No se pudo autenticar con el reloj. Verifica la Contraseña/Comm Key en Configuración."
            )
        if "timed out" in message.lower() or "timeout" in message.lower():
            raise HTTPException(status_code=408, detail=f"Tiempo de espera agotado al conectar con el reloj: {message}")
        if exc.__class__.__name__ == "ZKNetworkError":
            raise HTTPException(status_code=503, detail=f"Error de red al conectar con el reloj: {message}")
        raise HTTPException(status_code=503, detail=f"No se pudo conectar al reloj: {message}")


async def _flush_clock_flash(conn: Any):
    if hasattr(conn, "refresh_data"):
        await run_in_threadpool(conn.refresh_data)
    if hasattr(conn, "reg_event"):
        await run_in_threadpool(conn.reg_event)
    elif hasattr(conn, "reg_events"):
        await run_in_threadpool(conn.reg_events)


async def get_clock_connection(config: dict):
    if ZK is None:
        raise HTTPException(status_code=500, detail="Error: Librería pyzk no instalada en el servidor")

    device_ip = str(config.get("ip", "") or "").strip()
    if not device_ip:
        raise HTTPException(status_code=400, detail="Ingresa la IP del reloj antes de conectar.")

    device_port = _safe_parse_int(config.get("port", 4370), 4370)
    device_password = _safe_parse_int(config.get("password", 0), 0)
    force_udp = False

    last_error = None
    for attempt in range(1, 4):
        try:
            socket.setdefaulttimeout(10)
            print(f"Intentando conexión TCP al reloj {device_ip}:{device_port}....")
            zk_client = ZK(
                device_ip,
                port=device_port,
                timeout=10,
                password=device_password,
                force_udp=force_udp,
                ommit_ping=True
            )
            conn = await run_in_threadpool(zk_client.connect)
            await run_in_threadpool(conn.disable_device)
            return conn
        except Exception as exc:
            last_error = exc
            print(traceback.format_exc())
            print(f"[CLOCK_SYNC] Intento {attempt}/3 fallido: {exc}")
            if attempt < 3:
                await asyncio.sleep(2)

    error_text = str(last_error or "").lower()
    status_code = 408 if "timed out" in error_text or "timeout" in error_text else 503
    raise HTTPException(
        status_code=status_code,
        detail=f"No se pudo conectar al reloj ({device_ip}:{device_port}) tras 3 intentos. "
               f"Verifica red/puerto/comm key. Último error: {last_error}"
    )


def _serialize_clock_user(doc: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(doc)
    out["_id"] = str(out["_id"])
    return out


def _safe_parse_int(value: Any, default: int = 0) -> int:
    try:
        text = str(value).strip()
        if text == "":
            return default
        return int(text)
    except Exception:
        return default


def _resolve_clock_user_id(raw_user: Any) -> str:
    user_id = str(getattr(raw_user, "user_id", "") or "").strip()
    if user_id:
        return user_id
    return str(getattr(raw_user, "uid", "") or "").strip()


def _same_subnet(local_ip: str, clock_ip: str, prefix: int = 24) -> bool:
    try:
        local_net = ipaddress.ip_network(f"{local_ip}/{prefix}", strict=False)
        return ipaddress.ip_address(clock_ip) in local_net
    except Exception:
        return False


def _get_local_ipv4s() -> List[str]:
    ips = {"127.0.0.1"}
    try:
        hostname = socket.gethostname()
        for result in socket.getaddrinfo(hostname, None, socket.AF_INET):
            ip = result[4][0]
            if ip:
                ips.add(ip)
    except Exception:
        pass
    return sorted(ips)


@api_router.get("/clock/config")
async def get_clock_config(request: Request):
    await get_current_user(request)
    config = await db.clock_config.find_one({}, {"_id": 0})
    if not config:
        config = {
            "device_name": "",
            "ip": "",
            "port": 4370,
            "password": "",
            "rules": [
                {"name": "Turno General", "expected_entry_time": "09:00", "tolerance_minutes": 30}
            ],
            "connected": False,
            "last_sync": None,
        }
        await db.clock_config.insert_one(config)
    return config


@api_router.get("/clock/config/export")
async def export_clock_config(request: Request):
    await get_current_user(request)
    config = await db.clock_config.find_one({}, {"_id": 0}) or {}
    payload = json.dumps(config, ensure_ascii=False, default=str, indent=2).encode("utf-8")
    return StreamingResponse(
        io.BytesIO(payload),
        media_type="application/json; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=clock_config.json"},
    )


@api_router.put("/clock/config")
async def update_clock_config(payload: ClockConfigUpdate, request: Request):
    await get_current_user(request)
    data = payload.model_dump()
    data["updated_at"] = datetime.now(timezone.utc)
    await db.clock_config.update_one({}, {"$set": data}, upsert=True)
    return data


@api_router.post("/clock/config/import")
async def import_clock_config(payload: ClockConfigImportPayload, request: Request):
    await get_current_user(request)
    config_data = payload.model_dump()
    config_data["updated_at"] = datetime.now(timezone.utc)

    # Aplicar y validar los cambios de conexión de forma inmediata.
    conn = _connect_to_clock(config_data)
    try:
        conn.disable_device()
    finally:
        try:
            conn.enable_device()
        except Exception:
            pass
        try:
            conn.disconnect()
        except Exception:
            pass

    config_data["connected"] = True
    await db.clock_config.update_one({}, {"$set": config_data}, upsert=True)
    return {"message": "Configuración importada y aplicada", "config": config_data}


@api_router.post("/clock/test-connection")
async def test_clock_connection(request: Request):
    await get_current_user(request)
    config = await db.clock_config.find_one({}, {"_id": 0})
    if not config:
        raise HTTPException(status_code=400, detail="Configura primero el reloj checador")

    try:
        conn = _connect_to_clock(config)
    except HTTPException as exc:
        if config.get("ip") == "192.168.1.104":
            return {"status": "error", "message": "Reloj fuera de línea"}
        raise exc
    except Exception:
        if config.get("ip") == "192.168.1.104":
            return {"status": "error", "message": "Reloj fuera de línea"}
        raise
    try:
        return {"connected": True, "message": "Conexión y autenticación con reloj exitosas"}
    finally:
        try:
            conn.disconnect()
        except Exception:
            pass


@api_router.post("/clock/connection")
async def set_clock_connection(payload: ClockConnectionPayload, request: Request):
    await get_current_user(request)
    config = await db.clock_config.find_one({}, {"_id": 0})
    if not config:
        raise HTTPException(status_code=400, detail="Configura primero el reloj checador")

    if payload.connected:
        try:
            conn = _connect_to_clock(config)
        except HTTPException as exc:
            if config.get("ip") == "192.168.1.104":
                return {"status": "error", "message": "Reloj fuera de línea"}
            raise exc
        except Exception:
            if config.get("ip") == "192.168.1.104":
                return {"status": "error", "message": "Reloj fuera de línea"}
            raise
        try:
            pass
        finally:
            try:
                conn.disconnect()
            except Exception:
                pass

    await db.clock_config.update_one(
        {},
        {"$set": {"connected": payload.connected, "connection_changed_at": datetime.now(timezone.utc)}},
        upsert=True
    )
    return {"connected": payload.connected}


@api_router.get("/clock/status")
async def get_clock_status(request: Request):
    await get_current_user(request)
    config = await db.clock_config.find_one({}, {"_id": 0}) or {}
    users_count = await db.clock_users.count_documents({})
    last_event = await db.clock_events.find_one({}, sort=[("created_at", -1)])
    return {
        "connected": bool(config.get("connected", False)),
        "device_name": config.get("device_name", "Reloj Principal"),
        "ip": config.get("ip"),
        "port": config.get("port"),
        "users_count": users_count,
        "last_sync": config.get("last_sync"),
        "last_event_at": last_event.get("created_at") if last_event else None,
    }


@api_router.get("/clock/network-check")
async def get_clock_network_check(request: Request):
    await get_current_user(request)
    config = await db.clock_config.find_one({}, {"_id": 0}) or {}
    clock_ip = str(config.get("ip", "") or "").strip()
    local_ips = _get_local_ipv4s()
    same_subnet = any(_same_subnet(local_ip, clock_ip) for local_ip in local_ips if local_ip != "127.0.0.1") if clock_ip else False
    return {
        "clock_ip": clock_ip,
        "local_ips": local_ips,
        "same_subnet": same_subnet,
        "message": "Equipo y reloj en misma red" if same_subnet else "Equipo y reloj parecen estar en redes distintas"
    }


@api_router.get("/clock/device-info")
async def get_clock_device_info(request: Request):
    await get_current_user(request)
    config = await db.clock_config.find_one({}, {"_id": 0})
    if not config:
        raise HTTPException(status_code=400, detail="Configura primero el reloj checador")

    users_count = await db.clock_users.count_documents({})
    fingerprint_users = await db.clock_users.count_documents({"fingerprint_registered": True})
    face_users = await db.clock_users.count_documents({"face_registered": True})
    vein_users = await db.clock_users.count_documents({"vein_registered": True})

    info = {
        "device_name": config.get("device_name", "Reloj Principal"),
        "ip": config.get("ip"),
        "port": config.get("port"),
        "connected": bool(config.get("connected", False)),
        "users": users_count,
        "fingerprints": fingerprint_users,
        "faces": face_users,
        "veins": vein_users,
    }

    if info["connected"]:
        conn = None
        try:
            conn = _connect_to_clock(config)
            info["device_time"] = conn.get_time().isoformat() if hasattr(conn, "get_time") else None
        except Exception as exc:
            info["device_time_error"] = str(exc)
        finally:
            try:
                conn.disconnect()
            except Exception:
                pass

    return info


@api_router.get("/clock/users")
async def get_clock_users(request: Request):
    await get_current_user(request)
    users = await db.clock_users.find({}).sort("user_id", 1).to_list(2000)
    return [_serialize_clock_user(u) for u in users]


async def _sync_single_user_to_clock(user_doc: Dict[str, Any], mode: str = "manual") -> Dict[str, Any]:
    if mode != "wifi":
        return {"sync_mode": "manual", "sync_status": "pending"}

    config = await db.clock_config.find_one({}, {"_id": 0}) or {}
    if not config.get("connected"):
        return {"sync_mode": "manual", "sync_status": "pending", "message": "Reloj desconectado"}

    conn = _connect_to_clock(config)
    try:
        await run_in_threadpool(
            conn.set_user,
            uid=_safe_parse_int(user_doc.get("uid", user_doc.get("user_id")), 0),
            user_id=str(user_doc.get("user_id") or ""),
            name=str(user_doc.get("name") or ""),
            privilege=1 if user_doc.get("privilege") == "admin" else 0,
            password=str(user_doc.get("password") or ""),
            card=int(_safe_parse_int(user_doc.get("card_number"), 0)),
        )
        await _flush_clock_flash(conn)
        return {"sync_mode": "wifi", "sync_status": "synced", "message": "Sincronizado por WiFi"}
    finally:
        try:
            conn.disconnect()
        except Exception:
            pass


@api_router.post("/clock/users")
async def create_clock_user(payload: ClockUserBase, request: Request, sync_mode: Literal["wifi", "manual"] = "manual"):
    await get_current_user(request)
    normalized_user_id = payload.user_id.strip()
    if not normalized_user_id:
        raise HTTPException(status_code=400, detail="El ID del usuario no puede estar vacío.")

    exists = await db.clock_users.find_one({"user_id": normalized_user_id})
    if exists:
        raise HTTPException(status_code=409, detail="Ya existe un usuario con ese ID")
    doc = payload.model_dump()
    doc["user_id"] = normalized_user_id
    doc["uid"] = _safe_parse_int(normalized_user_id, 0)
    doc.update({
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc),
        "sync_status": "pending",
    })
    sync_result = await _sync_single_user_to_clock(doc, sync_mode)
    doc["sync_status"] = sync_result.get("sync_status", "pending")
    result = await db.clock_users.insert_one(doc)
    saved = await db.clock_users.find_one({"_id": result.inserted_id})
    return _serialize_clock_user(saved)


@api_router.put("/clock/users/{user_id}")
async def update_clock_user(user_id: str, payload: ClockUserUpdate, request: Request, sync_mode: Literal["wifi", "manual"] = "manual"):
    await get_current_user(request)
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No hay cambios para actualizar")
    updates["updated_at"] = datetime.now(timezone.utc)
    updates["sync_status"] = "pending"
    current_user = await db.clock_users.find_one({"user_id": user_id})
    if not current_user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    merged_user = {**current_user, **updates}
    sync_result = await _sync_single_user_to_clock(merged_user, sync_mode)
    updates["sync_status"] = sync_result.get("sync_status", "pending")
    result = await db.clock_users.update_one({"user_id": user_id}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    user = await db.clock_users.find_one({"user_id": user_id})
    return _serialize_clock_user(user)


@api_router.delete("/clock/users/{user_id}")
async def delete_clock_user(user_id: str, request: Request, sync_mode: Literal["wifi", "manual"] = "manual"):
    await get_current_user(request)
    if sync_mode == "wifi":
        config = await db.clock_config.find_one({}, {"_id": 0}) or {}
        if config.get("connected"):
            conn = _connect_to_clock(config)
            try:
                delete_fn = getattr(conn, "delete_user", None) or getattr(conn, "deleteUser", None)
                if delete_fn:
                    await run_in_threadpool(delete_fn, user_id)
                    await _flush_clock_flash(conn)
            finally:
                try:
                    conn.disconnect()
                except Exception:
                    pass
    result = await db.clock_users.delete_one({"user_id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return {"message": "Usuario eliminado"}


@api_router.post("/clock/users/pull")
async def pull_users_from_clock(request: Request):
    await get_current_user(request)
    config = await db.clock_config.find_one({}, {"_id": 0})
    if not config:
        raise HTTPException(status_code=400, detail="Configura primero el reloj checador")

    conn = _connect_to_clock(config)
    imported = 0
    try:
        conn.disable_device()
        users = conn.get_users() or []
        for user in users:
            user_id = _resolve_clock_user_id(user)
            if not user_id:
                continue
            privilege_value = _safe_parse_int(getattr(user, "privilege", 0), 0)
            doc = {
                "user_id": user_id,
                "name": getattr(user, "name", f"Empleado {user_id}") or f"Empleado {user_id}",
                "department": "General",
                "privilege": "admin" if privilege_value > 0 else "empleado",
                "password": getattr(user, "password", "") or "",
                "card_number": str(getattr(user, "card", "") or ""),
                "uid": _safe_parse_int(getattr(user, "uid", user_id), 0),
                "fingerprint_registered": True,
                "face_registered": False,
                "vein_registered": False,
                "work_schedule": "Turno General",
                "enabled": True,
                "sync_status": "synced",
                "updated_at": datetime.now(timezone.utc),
            }
            await db.clock_users.update_one(
                {"user_id": user_id},
                {"$set": doc, "$setOnInsert": {"created_at": datetime.now(timezone.utc)}},
                upsert=True
            )
            imported += 1
    finally:
        try:
            conn.enable_device()
        except Exception:
            pass
        try:
            conn.disconnect()
        except Exception:
            pass

    return {"imported": imported}


@api_router.post("/clock/users/push")
async def push_users_to_clock(request: Request):
    await get_current_user(request)
    config = await db.clock_config.find_one({}, {"_id": 0})
    if not config:
        raise HTTPException(status_code=400, detail="Configura primero el reloj checador")

    users = await db.clock_users.find({"sync_status": "pendiente"}).to_list(2000)
    conn = _connect_to_clock(config)
    pushed = 0
    errors = []

    try:
        try:
            conn.disable_device()
        except Exception as disable_exc:
            logger.warning("[CLOCK_USERS_PUSH] No se pudo deshabilitar el dispositivo antes del push: %s", disable_exc)
        for user in users:
            try:
                privilege = 14 if user.get("privilege") == "admin" else 0
                raw_uid = user.get("uid", user.get("user_id", ""))
                uid = _safe_parse_int(raw_uid, 0)
                if uid <= 0:
                    raise ValueError(
                        f"El usuario {user.get('user_id', '(sin ID)')} no tiene UID numérico válido. "
                        "Edita el ID de empleado para que sea numérico o vuelve a importarlo desde el reloj."
                    )
                card_value = _safe_parse_int(user.get("card_number", 0), 0)
                await run_in_threadpool(
                    conn.set_user,
                    uid=uid,
                    name=user.get("name", ""),
                    privilege=privilege,
                    password=user.get("password", ""),
                    user_id=str(user["user_id"]),
                    card=card_value
                )
                await _flush_clock_flash(conn)
                await db.clock_users.update_one(
                    {"_id": user["_id"]},
                    {"$set": {"sync_status": "sincronizado", "synced_at": datetime.now(timezone.utc), "last_error": None}}
                )
                pushed += 1
            except Exception as exc:
                logger.exception(
                    "[CLOCK_USERS_PUSH] Error al subir usuario al reloj | user_id=%s | uid=%s | nombre=%s | privilegio=%s | error=%s",
                    user.get("user_id"),
                    user.get("uid", user.get("user_id")),
                    user.get("name", ""),
                    user.get("privilege", "empleado"),
                    exc,
                )
                errors.append({"user_id": user.get("user_id"), "error": str(exc)})
                await db.clock_users.update_one(
                    {"_id": user["_id"]},
                    {"$set": {"sync_status": "error", "last_error": str(exc)}}
                )
    finally:
        try:
            await run_in_threadpool(conn.enable_device)
        except Exception:
            pass
        try:
            await run_in_threadpool(conn.disconnect)
        except Exception:
            pass

    return {"pushed": pushed, "errors": errors}


def _normalize_timestamp(raw_timestamp: Any) -> Optional[datetime]:
    if isinstance(raw_timestamp, datetime):
        return raw_timestamp
    if raw_timestamp is None:
        return None
    try:
        return datetime.fromisoformat(str(raw_timestamp))
    except Exception:
        return None


@api_router.post("/reports/sync")
async def sync_and_export_historical_report(request: Request):
    await get_current_user(request)
    config = await db.clock_config.find_one({}, {"_id": 0})
    if not config:
        raise HTTPException(status_code=400, detail="Configura primero el reloj checador")

    conn = None
    attendance = []
    clock_users = []
    try:
        conn = _connect_to_clock(config)
        conn.disable_device()
        attendance = await run_in_threadpool(conn.get_attendance)
        clock_users = await run_in_threadpool(conn.get_users)
    except HTTPException:
        raise
    except TimeoutError as exc:
        raise HTTPException(status_code=408, detail=f"Tiempo de espera agotado al sincronizar: {exc}")
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Error de hardware al sincronizar: {exc}")
    finally:
        if conn:
            try:
                await run_in_threadpool(conn.enable_device)
            except Exception:
                pass
            try:
                await run_in_threadpool(conn.disconnect)
            except Exception:
                pass

    user_name_map = {
        _resolve_clock_user_id(user): str(getattr(user, "name", "") or "").strip()
        for user in (clock_users or [])
        if _resolve_clock_user_id(user)
    }

    grouped_records: Dict[tuple, Dict[str, Any]] = {}
    synced_events: List[Dict[str, Any]] = []
    for item in attendance or []:
        user_id = str(getattr(item, "user_id", "") or "").strip()
        timestamp = getattr(item, "timestamp", None)
        ts = _normalize_timestamp(timestamp)
        if not user_id or not ts:
            continue

        user_name = user_name_map.get(user_id) or f"Empleado {user_id}"
        key = (user_id, ts.date().isoformat())
        current = grouped_records.get(key)
        if not current:
            current = {
                "id": user_id,
                "nombre": user_name,
                "depto": "General",
                "fecha": ts.date().isoformat(),
                "entrada": ts,
                "salida": ts,
            }
            grouped_records[key] = current
        else:
            if ts < current["entrada"]:
                current["entrada"] = ts
            if ts > current["salida"]:
                current["salida"] = ts

        synced_events.append({
            "clock_user_id": user_id,
            "employee_name": user_name,
            "timestamp": ts,
        })

    # Sync identidad de usuarios en MongoDB
    for row in grouped_records.values():
        await db.clock_users.update_one(
            {"user_id": row["id"]},
            {
                "$set": {
                    "name": row["nombre"],
                    "department": row["depto"],
                    "sync_status": "sincronizado",
                    "updated_at": datetime.now(timezone.utc),
                },
                "$setOnInsert": {
                    "created_at": datetime.now(timezone.utc),
                    "uid": _safe_parse_int(row["id"], 0),
                    "privilege": "empleado",
                    "password": "",
                    "card_number": "",
                    "fingerprint_registered": True,
                    "face_registered": False,
                    "vein_registered": False,
                    "work_schedule": "Turno General",
                    "enabled": True,
                },
            },
            upsert=True,
        )

    if synced_events:
        await db.clock_events.insert_one({
            "events": synced_events,
            "total_events": len(synced_events),
            "created_at": datetime.now(timezone.utc),
        })

    await db.clock_config.update_one(
        {},
        {"$set": {"last_sync": datetime.now(timezone.utc), "connected": True}},
        upsert=True,
    )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Nombre", "Depto", "Fecha", "Entrada", "Salida"])
    for row in sorted(grouped_records.values(), key=lambda r: (r["fecha"], r["id"])):
        writer.writerow([
            row["id"],
            row["nombre"],
            row["depto"],
            row["fecha"],
            row["entrada"].strftime("%H:%M:%S"),
            row["salida"].strftime("%H:%M:%S"),
        ])

    csv_bytes = output.getvalue().encode("utf-8-sig")
    output.close()
    filename = f"Reporte_de_Asistencia_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@api_router.get("/clock/settings")
async def get_clock_att_settings(request: Request):
    await get_current_user(request)
    config = await db.clock_config.find_one({}, {"_id": 0})
    if not config:
        raise HTTPException(status_code=400, detail="Configura primero el reloj checador")

    conn = None
    try:
        conn = _connect_to_clock(config)
        conn.disable_device()
        workcodes = []
        if hasattr(conn, "get_workcode"):
            workcodes = await run_in_threadpool(conn.get_workcode)
    except HTTPException:
        raise
    except TimeoutError as exc:
        raise HTTPException(status_code=408, detail=f"Tiempo de espera agotado al leer horarios: {exc}")
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Error de hardware al leer horarios: {exc}")
    finally:
        if conn:
            try:
                await run_in_threadpool(conn.enable_device)
            except Exception:
                pass
            try:
                await run_in_threadpool(conn.disconnect)
            except Exception:
                pass

    rows = []
    for idx, code in enumerate(workcodes or [], start=1):
        rows.append({
            "numero": idx,
            "entrada": str(getattr(code, "start", "09:00") or "09:00"),
            "salida": str(getattr(code, "end", "18:00") or "18:00"),
            "tiempo_extra": _safe_parse_int(getattr(code, "ot", 0), 0),
        })

    if not rows:
        stored = await db.clock_settings.find_one({}, {"_id": 0})
        rows = (stored or {}).get("settings", [])
    return {"settings": rows}


@api_router.post("/clock/settings")
async def save_clock_att_settings(payload: AttSettingsPayload, request: Request):
    await get_current_user(request)
    config = await db.clock_config.find_one({}, {"_id": 0})
    if not config:
        raise HTTPException(status_code=400, detail="Configura primero el reloj checador")

    conn = None
    try:
        conn = _connect_to_clock(config)
        conn.disable_device()
        for row in payload.settings:
            if hasattr(conn, "set_workcode"):
                try:
                    await run_in_threadpool(conn.set_workcode, int(row.numero), row.entrada)
                except TypeError:
                    await run_in_threadpool(conn.set_workcode, int(row.numero))
                await _flush_clock_flash(conn)
    except HTTPException:
        raise
    except TimeoutError as exc:
        raise HTTPException(status_code=408, detail=f"Tiempo de espera agotado al guardar horarios: {exc}")
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Error de hardware al guardar horarios: {exc}")
    finally:
        if conn:
            try:
                await run_in_threadpool(conn.enable_device)
            except Exception:
                pass
            try:
                await run_in_threadpool(conn.disconnect)
            except Exception:
                pass

    await db.clock_settings.update_one(
        {},
        {"$set": {"settings": [row.model_dump() for row in payload.settings], "updated_at": datetime.now(timezone.utc)}},
        upsert=True,
    )
    return {"message": "Horarios guardados en reloj", "settings": payload.settings}


async def sync_clock_data(config: Dict[str, Any]):
    conn = None
    attendance = []
    detected_records = 0
    skipped_duplicates = 0
    inserted_records = 0

    rules = config.get("rules", []) or []
    default_rule = next((r for r in rules if str(r.get("name", "")).strip().lower() == "turno general"), None)
    if not default_rule and rules:
        default_rule = rules[0]
    if not default_rule:
        default_rule = {"name": "Turno General", "expected_entry_time": "09:00", "tolerance_minutes": 30}

    rule_by_name = {str(r.get("name", "")).strip(): r for r in rules if str(r.get("name", "")).strip()}
    clock_users = await db.clock_users.find({}, {"user_id": 1, "work_schedule": 1}).to_list(5000)
    schedule_by_clock_user = {
        str(u.get("user_id", "")).strip(): str(u.get("work_schedule", "Turno General") or "Turno General").strip()
        for u in clock_users
        if str(u.get("user_id", "")).strip()
    }

    try:
        print(f"Intentando conectar a IP: {config.get('ip')}:{config.get('port', 4370)}")
        conn = await get_clock_connection(config)
        print("Conexión exitosa")
        await run_in_threadpool(conn.disable_device)
        attendance = await run_in_threadpool(conn.get_attendance)
        detected_records = len(attendance or [])
        print(f"Registros encontrados: {detected_records}")
        print(f"[CLOCK_SYNC] Registros detectados por zk.get_attendance(): {detected_records}")
        logger.info("Registros detectados por zk.get_attendance(): %s", detected_records)
    except HTTPException:
        raise
    except TimeoutError as exc:
        raise HTTPException(status_code=408, detail=f"Tiempo de espera agotado al sincronizar asistencias: {exc}")
    except Exception as exc:
        if exc.__class__.__name__ == "ZKNetworkError":
            raise HTTPException(status_code=503, detail=f"Error de red con el reloj: {exc}")
        raise HTTPException(status_code=503, detail=f"No se pudo sincronizar con el reloj: {exc}")
    finally:
        if conn:
            try:
                await run_in_threadpool(conn.enable_device)
            except Exception:
                pass
            try:
                await run_in_threadpool(conn.disconnect)
            except Exception:
                pass

    for item in attendance or []:
        timestamp = getattr(item, "timestamp", None)
        clock_user_id = str(getattr(item, "user_id", "") or "").strip()
        source_device_id = str(getattr(item, "device_id", "") or "").strip()
        if not clock_user_id or not isinstance(timestamp, datetime):
            continue

        normalized_ts = timestamp.replace(tzinfo=None) if timestamp.tzinfo else timestamp
        id_candidates: List[Any] = [clock_user_id]
        if clock_user_id.isdigit():
            id_candidates.append(int(clock_user_id))
        employee_doc = await db.employees.find_one(
            {"employee_id": {"$in": id_candidates}},
            {"employee_id": 1, "name": 1}
        )
        if not employee_doc:
            employee_id = "ID Desconocido"
            employee_name = f"ID Reloj: {clock_user_id}"
        else:
            employee_id = str(employee_doc.get("employee_id", "") or "").strip() or "ID Desconocido"
            employee_name = str(employee_doc.get("name", "") or "").strip() or f"ID Reloj: {clock_user_id}"

        schedule_name = schedule_by_clock_user.get(clock_user_id, "Turno General")
        rule = rule_by_name.get(schedule_name) or default_rule
        expected_entry_time = str(rule.get("expected_entry_time", "09:00") or "09:00")
        tolerance_minutes = _safe_parse_int(rule.get("tolerance_minutes", 30), 30)
        try:
            expected_hour, expected_minute = [int(part) for part in expected_entry_time.split(":", 1)]
        except Exception:
            expected_hour, expected_minute = 9, 0

        scheduled_dt = normalized_ts.replace(hour=expected_hour, minute=expected_minute, second=0, microsecond=0)
        delay_minutes = max((normalized_ts - scheduled_dt).total_seconds() / 60.0, 0.0)
        status = "RETARDO" if delay_minutes > tolerance_minutes else "PRESENTE"

        record = {
            "employee_id": employee_id,
            "employee_name": employee_name,
            "timestamp": normalized_ts,
            "type": str(getattr(item, "status", getattr(item, "punch", 0))),
            "clock_user_id": clock_user_id,
            "source_device_id": source_device_id,
            "work_schedule": schedule_name,
            "expected_entry_time": expected_entry_time,
            "tolerance_minutes": tolerance_minutes,
            "delay_minutes": round(delay_minutes, 2),
            "status": status,
        }

        exists = await db.attendance_records.find_one(
            {"employee_id": record["employee_id"], "timestamp": record["timestamp"]},
            {"_id": 1}
        )
        if exists:
            skipped_duplicates += 1
            continue
        await db.attendance_records.insert_one(record)
        inserted_records += 1

    await db.clock_config.update_one(
        {},
        {"$set": {"connected": True, "last_sync": datetime.now()}},
        upsert=True
    )
    return detected_records, skipped_duplicates, inserted_records


@api_router.get("/clock/attendance/live")
async def get_live_attendance(request: Request, limit: int = 100):
    await get_current_user(request)
    docs = await db.clock_events.find({}, {"events": 1, "created_at": 1}).sort("created_at", -1).to_list(5)
    events = []
    for doc in docs:
        for event in doc.get("events", []):
            timestamp = event.get("timestamp")
            if isinstance(timestamp, datetime):
                ts_iso = timestamp.isoformat()
            else:
                ts_iso = str(timestamp)
            events.append({
                "employee_id": event.get("employee_id"),
                "timestamp": ts_iso,
                "received_at": doc.get("created_at").isoformat() if isinstance(doc.get("created_at"), datetime) else str(doc.get("created_at")),
            })
    events = sorted(events, key=lambda e: e["timestamp"], reverse=True)[: max(1, min(limit, 500))]
    return {"events": events}


@api_router.post("/clock/sync", response_model=ClockSyncResponse)
async def sync_clock_attendance(request: Request):
    current_user = await get_current_user(request)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo un administrador puede sincronizar el reloj.")

    config = await db.clock_config.find_one({}, {"_id": 0})
    if not config:
        raise HTTPException(status_code=400, detail="Configura primero el reloj checador")

    detected_records, skipped_duplicates, inserted_records = await sync_clock_data(config)

    return ClockSyncResponse(
        synced_records=inserted_records,
        generated_report_id=None,
        message="Sincronización completada desde el reloj checador",
        detected_records=detected_records,
        skipped_duplicates=skipped_duplicates,
        inserted_records=inserted_records,
    )


@api_router.post("/clock/users/import")
async def import_clock_users(request: Request):
    current_user = await get_current_user(request)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo un administrador puede importar usuarios del reloj.")

    config = await db.clock_config.find_one({}, {"_id": 0})
    if not config:
        raise HTTPException(status_code=400, detail="Configura primero el reloj checador")

    conn = None
    imported = 0
    skipped = 0
    try:
        print(f"Intentando conectar a IP: {config.get('ip')}:{config.get('port', 4370)}")
        conn = await get_clock_connection(config)
        print("Conexión exitosa")
        await run_in_threadpool(conn.disable_device)
        clock_users = await run_in_threadpool(conn.get_users)

        for user in clock_users or []:
            clock_user_id = _resolve_clock_user_id(user)
            if not clock_user_id:
                skipped += 1
                continue

            id_candidates: List[Any] = [clock_user_id]
            if clock_user_id.isdigit():
                id_candidates.append(int(clock_user_id))

            existing = await db.employees.find_one({"internal_clock_id": {"$in": id_candidates}}, {"_id": 1})
            if existing:
                skipped += 1
                continue

            await db.employees.insert_one({
                "employee_id": clock_user_id,
                "name": getattr(user, "name", f"Empleado {clock_user_id}") or f"Empleado {clock_user_id}",
                "department": "Reloj checador",
                "internal_clock_id": clock_user_id,
                "created_at": datetime.now(timezone.utc),
                "updated_at": datetime.now(timezone.utc),
            })
            imported += 1
    except HTTPException:
        raise
    except TimeoutError as exc:
        raise HTTPException(status_code=408, detail=f"Tiempo de espera agotado al importar usuarios del reloj: {exc}")
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"No se pudieron importar usuarios del reloj: {exc}")
    finally:
        if conn:
            try:
                await run_in_threadpool(conn.enable_device)
            except Exception:
                pass
            try:
                await run_in_threadpool(conn.disconnect)
            except Exception:
                pass

    return {"imported": imported, "skipped": skipped}


@api_router.get("/clock/events")
async def get_clock_events(request: Request, limit: int = 10):
    await get_current_user(request)
    docs = await db.clock_events.find({}, {"events": 0}).sort("created_at", -1).to_list(max(1, min(limit, 100)))
    for doc in docs:
        doc["_id"] = str(doc["_id"])
    return docs


# Root route
@api_router.get("/")
async def root():
    return {"message": "Sistema de Control de Asistencia API", "version": os.environ.get("APP_VERSION", "1.0.0")}

# Include the router in the main app
app.include_router(api_router)

def get_allowed_origins() -> List[str]:
    raw_origins = os.environ.get("FRONTEND_URL", "http://localhost:3000")
    origins = [origin.strip() for origin in raw_origins.split(",") if origin.strip()]

    expanded_origins = set(origins)
    for origin in origins:
        parsed = urlparse(origin)
        if not parsed.scheme or not parsed.netloc:
            continue
        if "localhost" in parsed.netloc:
            expanded_origins.add(origin.replace("localhost", "127.0.0.1"))
        if "127.0.0.1" in parsed.netloc:
            expanded_origins.add(origin.replace("127.0.0.1", "localhost"))

    # Common local dev origins
    expanded_origins.update({
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        
    })
    return sorted(expanded_origins)

ALLOWED_ORIGINS = get_allowed_origins()

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.middleware("http")
async def global_exception_trace_middleware(request: Request, call_next):
    try:
        return await call_next(request)
    except Exception:
        traceback.print_exc()
        raise

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Startup event
@app.on_event("startup")
async def startup_event():
    logger.info(f"CORS allow_origins: {ALLOWED_ORIGINS}")

    async def upsert_seed_user(email: str, password: str, name: str, role: str, extra_fields: Optional[Dict[str, Any]] = None):
        existing = await db.users.find_one({"email": email})
        extra_fields = extra_fields or {}
        base_updates = {
            "name": name,
            "role": role,
            **extra_fields
        }

        if existing is None:
            await db.users.insert_one({
                "email": email,
                "password_hash": hash_password(password),
                "created_at": datetime.now(timezone.utc),
                **base_updates
            })
            logger.info(f"Seed user created: {email}")
            return

        updates = dict(base_updates)
        if not verify_password(password, existing["password_hash"]):
            updates["password_hash"] = hash_password(password)
            logger.info(f"Seed user password updated: {email}")
        await db.users.update_one({"email": email}, {"$set": updates})

    # Seed users
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@example.com")
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    await upsert_seed_user(admin_email, admin_password, "Administrador", "admin")

    developer_email = os.environ.get("DEVELOPER_EMAIL", "leijahector5@gmail.com")
    developer_password = os.environ.get("DEVELOPER_PASSWORD", "/Leija091105")
    await upsert_seed_user(
        developer_email,
        developer_password,
        "Héctor Leija",
        "developer",
        {
            "developer_options": {
                "feature_flags": {
                    "beta_dashboard": True,
                    "diagnostic_mode": True,
                    "export_debug_logs": True
                },
                "can_manage_clock_config": True,
                "can_trigger_manual_sync": True
            }
        }
    )
    
    # Create indexes
    await db.users.create_index("email", unique=True)
    await db.reports.create_index("upload_date")
    await db.clock_users.create_index("user_id", unique=True)
    
    # Initialize settings if not exists
    settings = await db.settings.find_one({})
    if not settings:
        await db.settings.insert_one({
            "entry_time": "09:00",
            "tolerance_minutes": 30,
            "work_hours": 9
        })

    clock_config = await db.clock_config.find_one({})
    if not clock_config:
        await db.clock_config.insert_one({
            "device_name": "",
            "ip": "",
            "port": 4370,
            "password": "",
            "rules": [
                {"name": "Turno General", "expected_entry_time": "09:00", "tolerance_minutes": 30}
            ],
            "connected": False,
            "last_sync": None,
            "created_at": datetime.now(timezone.utc)
        })
    await db.clock_events.create_index("created_at")
    
    # Write test credentials
    credentials_path = Path("/app/memory/test_credentials.md")
    credentials_path.parent.mkdir(parents=True, exist_ok=True)
    credentials_path.write_text(f"""# Test Credentials

## Admin Account
- Email: {admin_email}
- Password: {admin_password}
- Role: admin

## Developer Account
- Email: {developer_email}
- Password: {developer_password}
- Role: developer

## Auth Endpoints
- POST /api/auth/login
- POST /api/auth/logout
- GET /api/auth/me
- POST /api/auth/refresh
""")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
